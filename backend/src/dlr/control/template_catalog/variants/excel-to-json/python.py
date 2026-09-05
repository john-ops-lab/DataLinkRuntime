"""Bounded XLSX/XLS data reader that never evaluates formulas or active content."""

from __future__ import annotations

import base64
import json
import math
import re
from collections.abc import Callable, Iterator
from contextlib import suppress
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from zipfile import BadZipFile, ZipFile

# Excel 转 JSON：可修改的配置集中在这里。
# 运行时提供待处理的数据或文件；处理规则在下面配置。
# 调试时可传入 JSON 对象覆盖同名配置；嵌套对象需要完整填写。
CONFIG = {
    # 工作表名称；null 使用第一个工作表。
    "sheet": None,
    # 需要读取的单元格范围，例如 A1:D100。
    "range": "A1:D100",
    # 是否把首行作为字段名。
    "header": True,
    # 作为表头的行号，从 1 开始。
    "header_row": 1,
    # 空单元格处理：null 保留空值、empty-string 转为空字符串、omit 忽略空字段。
    "null_policy": "null",
    # 单个文件读取大小上限，单位字节。
    "max_file_bytes": 8388608,
    # 最多读取的行数。
    "max_rows": 5000,
    # 最多读取的列数。
    "max_columns": 200,
    # 返回结果大小上限，单位字节。
    "max_output_bytes": 4194304,
}


_A1_RANGE = re.compile(r"^([A-Z]+)([1-9][0-9]*):([A-Z]+)([1-9][0-9]*)$", re.IGNORECASE)
_ACTIVE_MEMBER = re.compile(
    r"(?:^|/)(?:vbaproject\.bin|activex/|embeddings/|externallinks/|connections\.xml|"
    r"macrosheets/|dialogsheets/|customui/)",
    re.IGNORECASE,
)
_STABLE_ERRORS = frozenset(
    {
        "input_must_be_object",
        "workbook_input_file_required",
        "invalid_null_policy",
        "workbook_too_large",
        "macro_enabled_workbook_rejected",
        "unsupported_workbook_format",
        "workbook_archive_limit",
        "encrypted_workbook_rejected",
        "workbook_external_links_rejected",
        "workbook_active_content_rejected",
        "workbook_macros_rejected",
        "invalid_xlsx_package",
        "unsupported_xlsx_compression",
        "sheet_not_found",
        "invalid_range",
        "range_start_outside_selection",
        "header_row_outside_range",
        "invalid_or_duplicate_header",
        "unsupported_cell_type",
    }
)


def _positive(value: object, default: int, maximum: int) -> int:
    return (
        min(value, maximum)
        if isinstance(value, int) and not isinstance(value, bool) and value > 0
        else default
    )


def _cell(value: object) -> object:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("unsupported_cell_type")
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="milliseconds")
    if isinstance(value, time):
        return value.isoformat(timespec="milliseconds")
    if isinstance(value, date):
        return f"{value.isoformat()}T00:00:00.000"
    if isinstance(value, (bytes, bytearray, memoryview)):
        return {"$binary_base64": base64.b64encode(bytes(value)).decode("ascii")}
    raise ValueError("unsupported_cell_type")


def _column_number(value: str) -> int:
    result = 0
    for character in value.upper():
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def _bounds(
    input: dict[str, object],
    last_row: int,
    last_column: int,
    max_columns: int,
) -> tuple[int, int, int, int, bool]:
    raw_range = input.get("range")
    if raw_range is not None:
        if not isinstance(raw_range, str) or not (match := _A1_RANGE.fullmatch(raw_range)):
            raise ValueError("invalid_range")
        start_column = _column_number(match.group(1))
        start_row = int(match.group(2))
        desired_end_column = _column_number(match.group(3))
        desired_end_row = int(match.group(4))
        if (
            start_row > desired_end_row
            or start_column > desired_end_column
            or desired_end_row > 1_048_576
            or desired_end_column > 16_384
        ):
            raise ValueError("invalid_range")
    else:
        start_row = _positive(input.get("start_row"), 1, 1_048_576)
        start_column = _positive(input.get("start_column"), 1, 16_384)
        desired_end_row = max(start_row, last_row)
        desired_end_column = max(start_column, last_column)
    start_row = _positive(input.get("start_row"), start_row, 1_048_576)
    start_column = _positive(input.get("start_column"), start_column, 16_384)
    if start_row > desired_end_row or start_column > desired_end_column:
        raise ValueError("range_start_outside_selection")
    if input.get("header", True) and input.get("header_row") is not None:
        header_row = _positive(input.get("header_row"), start_row, 1_048_576)
        if header_row < start_row or header_row > desired_end_row:
            raise ValueError("header_row_outside_range")
        start_row = header_row
    end_column = min(desired_end_column, start_column + max_columns - 1)
    return (
        start_row,
        start_column,
        desired_end_row,
        end_column,
        desired_end_column > end_column,
    )


@dataclass
class _RowSource:
    sheets: list[str]
    rows: Iterator[tuple[int, list[object], bool]]
    start_row: int
    start_column: int
    end_column: int
    column_limited: bool
    close: Callable[[], None]
    formulas: bool = False


def _apply_null_policy(value: object, policy: str) -> object:
    return "" if value is None and policy == "empty-string" else value


def _inspect_xlsx(path: Path) -> None:
    try:
        with ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > 10_000:
                raise ValueError("workbook_archive_limit")
            expanded = 0
            relationship_bytes = 0
            for member in members:
                name = member.filename.replace("\\", "/")
                expanded += member.file_size
                if member.flag_bits & 1:
                    raise ValueError("encrypted_workbook_rejected")
                if member.file_size < 0 or member.file_size > 67_108_864 or expanded > 134_217_728:
                    raise ValueError("workbook_archive_limit")
                if _ACTIVE_MEMBER.search(name):
                    if "externallinks/" in name.casefold() or name.casefold().endswith(
                        "connections.xml"
                    ):
                        raise ValueError("workbook_external_links_rejected")
                    raise ValueError("workbook_active_content_rejected")
                lowered_name = name.casefold()
                if lowered_name.endswith(".rels") or lowered_name == "[content_types].xml":
                    with archive.open(member) as stream:
                        content = stream.read(1_048_577)
                    relationship_bytes += len(content)
                    if len(content) > 1_048_576 or relationship_bytes > 4_194_304:
                        raise ValueError("workbook_archive_limit")
                    lowered = content.lower()
                    if re.search(rb"targetmode\s*=\s*['\"]external", lowered):
                        raise ValueError("workbook_external_links_rejected")
                    if any(
                        marker in lowered
                        for marker in (b"vbaproject", b"macroenabled", b"activex", b"oleobject")
                    ):
                        raise ValueError("workbook_active_content_rejected")
    except BadZipFile:
        raise ValueError("invalid_xlsx_package") from None


def _xlsx(path: Path, input: dict[str, object], _max_rows: int, max_columns: int) -> _RowSource:
    import openpyxl

    _inspect_xlsx(path)
    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )
    try:
        names = list(workbook.sheetnames)
        selected = input.get("sheet")
        if not names or (isinstance(selected, str) and selected and selected not in names):
            raise ValueError("sheet_not_found")
        sheet = workbook[selected] if isinstance(selected, str) and selected else workbook[names[0]]
        minimum_row, minimum_column, selection_end_row, end_column, column_limited = _bounds(
            input, sheet.max_row, sheet.max_column, max_columns
        )
    except Exception:
        with suppress(Exception):
            workbook.close()
        raise

    def rows() -> Iterator[tuple[int, list[object], bool]]:
        maximum_row = min(selection_end_row, sheet.max_row)
        if maximum_row < minimum_row:
            return
        for row_at, raw in enumerate(
            sheet.iter_rows(
                min_row=minimum_row,
                min_col=minimum_column,
                max_row=maximum_row,
                max_col=end_column,
            ),
            start=minimum_row,
        ):
            values: list[object] = []
            formulas = False
            for cell in raw:
                if cell.data_type == "f":
                    formulas = True
                    values.append(None)
                else:
                    values.append(_cell(cell.value))
            yield row_at, values, formulas

    return _RowSource(
        sheets=names,
        rows=rows(),
        start_row=minimum_row,
        start_column=minimum_column,
        end_column=end_column,
        column_limited=column_limited,
        close=lambda: workbook.close(),
    )


def _xls(path: Path, input: dict[str, object], _max_rows: int, max_columns: int) -> _RowSource:
    """Read stored BIFF cell values without invoking a formula or macro evaluator."""
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True, formatting_info=False)
    try:
        names = workbook.sheet_names()
        selected = input.get("sheet")
        if not names or (isinstance(selected, str) and selected and selected not in names):
            raise ValueError("sheet_not_found")
        sheet = (
            workbook.sheet_by_name(selected)
            if isinstance(selected, str) and selected
            else workbook.sheet_by_index(0)
        )
        minimum_row, minimum_column, selection_end_row, end_column, column_limited = _bounds(
            input,
            max(1, sheet.nrows),
            max(1, sheet.ncols),
            max_columns,
        )
    except Exception:
        with suppress(Exception):
            workbook.release_resources()
        raise

    def rows() -> Iterator[tuple[int, list[object], bool]]:
        for row_at in range(minimum_row - 1, min(sheet.nrows, selection_end_row)):
            values: list[object] = []
            for column_at in range(minimum_column - 1, end_column):
                if column_at >= sheet.ncols:
                    values.append(None)
                    continue
                cell = sheet.cell(row_at, column_at)
                if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
                    values.append(None)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    values.append(bool(cell.value))
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(_cell(xlrd.xldate_as_datetime(cell.value, workbook.datemode)))
                else:
                    values.append(_cell(cell.value))
            yield row_at + 1, values, False

    return _RowSource(
        sheets=names,
        rows=rows(),
        start_row=minimum_row,
        start_column=minimum_column,
        end_column=end_column,
        column_limited=column_limited,
        close=workbook.release_resources,
    )


def _legacy_source(
    value: tuple[list[str], list[list[object]], bool, bool],
    input: dict[str, object],
) -> _RowSource:
    names, raw_rows, formulas, column_limited = value
    start_row = _positive(input.get("start_row"), 1, 1_048_576)
    start_column = _positive(input.get("start_column"), 1, 16_384)
    raw_range = input.get("range")
    if isinstance(raw_range, str) and (match := _A1_RANGE.fullmatch(raw_range)):
        start_row = _positive(input.get("start_row"), int(match.group(2)), 1_048_576)
        start_column = _positive(input.get("start_column"), _column_number(match.group(1)), 16_384)
    if input.get("header", True) and input.get("header_row") is not None:
        start_row = _positive(input.get("header_row"), start_row, 1_048_576)
    width = max((len(row) for row in raw_rows), default=1)
    rows = (
        (start_row + at, [_cell(cell) for cell in raw], False) for at, raw in enumerate(raw_rows)
    )
    return _RowSource(
        sheets=names,
        rows=rows,
        start_row=start_row,
        start_column=start_column,
        end_column=start_column + width - 1,
        column_limited=column_limited,
        close=lambda: None,
        formulas=formulas,
    )


def _checkpoint(
    reasons: list[str], next_row: int | None, next_column: int | None
) -> dict[str, object] | None:
    if not reasons:
        return None
    result: dict[str, object] = {"reason": reasons[0] if len(reasons) == 1 else "multiple_limits"}
    if len(reasons) > 1:
        result["limits"] = reasons
    if next_row is not None:
        result["next_row"] = next_row
    if next_column is not None:
        result["next_column"] = next_column
    return result


def _handle(context, input):
    if not isinstance(input, dict):
        raise ValueError("input_must_be_object")
    if not context.input_files:
        raise ValueError("workbook_input_file_required")
    item = context.input_files[0]
    max_file_bytes = _positive(input.get("max_file_bytes"), 8_388_608, 33_554_432)
    max_rows = _positive(input.get("max_rows"), 5_000, 50_000)
    max_columns = _positive(input.get("max_columns"), 200, 2_000)
    max_output_bytes = _positive(input.get("max_output_bytes"), 4_194_304, 16_777_216)
    null_policy = input.get("null_policy", "null")
    if null_policy not in {"null", "empty-string", "omit"}:
        raise ValueError("invalid_null_policy")
    if item.size_bytes > max_file_bytes or Path(item.path).stat().st_size > max_file_bytes:
        raise ValueError("workbook_too_large")
    suffix = Path(item.original_name).suffix.lower()
    if suffix in {".xlsm", ".xltm", ".xlam"}:
        raise ValueError("macro_enabled_workbook_rejected")
    if suffix == ".xlsx":
        loaded = _xlsx(item.path, input, max_rows, max_columns)
    elif suffix == ".xls":
        loaded = _xls(item.path, input, max_rows, max_columns)
    else:
        raise ValueError("unsupported_workbook_format")
    source = loaded if isinstance(loaded, _RowSource) else _legacy_source(loaded, input)
    rows: list[object] = []
    output_bytes = 2
    reasons = ["column_limit"] if source.column_limited else []
    next_column = source.end_column + 1 if source.column_limited else None
    next_row: int | None = None
    formulas = source.formulas
    headers: list[str] | None = None
    iterator = iter(source.rows)
    try:
        if input.get("header", True):
            try:
                _, raw_header, header_formulas = next(iterator)
            except StopIteration:
                raw_header = []
                header_formulas = False
            formulas = formulas or header_formulas
            if raw_header:
                headers = []
                for at, value in enumerate(raw_header):
                    if isinstance(value, (dict, list)):
                        raise ValueError("invalid_or_duplicate_header")
                    headers.append(str(value) if value not in (None, "") else f"column_{at + 1}")
                if len(headers) != len(set(headers)):
                    raise ValueError("invalid_or_duplicate_header")

        for physical_row, raw, row_formulas in iterator:
            formulas = formulas or row_formulas
            if len(rows) >= max_rows:
                reasons.insert(0, "row_limit")
                next_row = physical_row
                break
            if headers is not None:
                padded = raw[: len(headers)] + [None] * max(0, len(headers) - len(raw))
                value = {
                    name: _apply_null_policy(padded[at], null_policy)
                    for at, name in enumerate(headers)
                    if null_policy != "omit" or padded[at] is not None
                }
            else:
                value = [_apply_null_policy(cell, null_policy) for cell in raw]
            encoded_bytes = len(
                json.dumps(
                    value,
                    ensure_ascii=False,
                    allow_nan=False,
                    separators=(",", ":"),
                ).encode()
            ) + (1 if rows else 0)
            if output_bytes + encoded_bytes > max_output_bytes:
                reasons.insert(0, "output_limit")
                next_row = physical_row
                break
            rows.append(value)
            output_bytes += encoded_bytes
    finally:
        with suppress(Exception):
            source.close()
    checkpoint = _checkpoint(reasons, next_row, next_column)
    return {
        "sheets": source.sheets[:100],
        "rows": rows,
        "count": len(rows),
        "partial": checkpoint is not None,
        "checkpoint": checkpoint,
        "active_content": {
            "executed": False,
            "formulas_replaced_with_null": formulas,
            "legacy_xls_data_only": suffix == ".xls",
            "ooxml_preflight": suffix == ".xlsx",
        },
    }


def handle(context, input):
    if input is None:
        input = {}
    if not isinstance(input, dict):
        raise ValueError("输入必须是 JSON 对象")
    input = {**CONFIG, **input}
    try:
        return _handle(context, input)
    except ValueError as error:
        code = str(error)
        if code in _STABLE_ERRORS:
            raise ValueError(code) from None
        raise ValueError("excel_operation_failed") from None
    except Exception:
        raise ValueError("excel_operation_failed") from None
